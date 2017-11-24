#!/usr/bin/env python3
from openpyxl import load_workbook
import json
import argparse
import traceback
import sys
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
from ansicolor import black, red, blue
from AutoTranslateCommon import transmap_filename, to_xlsx
from AutoTranslationTranslator import FullAutoTranslator

def get_transmap(filename):
    wb = load_workbook(filename=filename)
    sheet = wb[wb.sheetnames[0]]
    tmap = {}
    for row in sheet.rows:
        if row[0].value == "Count": continue # Header
        count = int(row[0].value)
        utr_count = int(row[1].value)
        engl = row[2]
        transl = row[3]
        tmap[str(engl.value)] = str(transl.value)
    return [{"english": engl, "translated": transl, "count": count, "untranslated_count": utr_count}
            for engl,transl in tmap.items()]

def _translate(entry, translator, force=False, tries_left=5):
    engl = entry["english"]
    if not force and entry["translated"] is not None and entry["translated"] == "":
        return entry # leave as is
    try:
        transl = translator.translate(engl)
    except:
        print(black("Autotranslate fail for string '{}'".format(engl), bold=True))
        traceback.print_exception(*sys.exc_info())
        return entry
    entry["translated"] = transl
    #print("{} ==> {}".format(engl, entry["translated"]))
    return entry


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-l','--language', help='The language')
    parser.add_argument('-j','--parallel', type=int, default=16, help='The language')
    parser.add_argument('iftags', help='The IF tags XLSX file')
    parser.add_argument('texttags', help='The text tags XLSX file')
    args = parser.parse_args()

    iftags = get_transmap(args.iftags)
    texttags = get_transmap(args.texttags)
    print("Found {} iftags".format(len(iftags)))
    print("Found {} text tags".format(len(texttags)))

    executor = ThreadPoolExecutor(args.parallel)

    # Translate them
    fat = FullAutoTranslator(args.language, limit=1000000)
    #
    # Process IF tags
    #
    _futures1 = [executor.submit(_translate, entry, translator=fat) for entry in iftags]
    _futures2 = [executor.submit(_translate, entry, translator=fat) for entry in texttags]
    _futures = _futures1 + _futures2
    kwargs = {
        'total': len(_futures),
        'unit': 'it',
        'unit_scale': True,
        'leave': True
    }
    for future in tqdm(concurrent.futures.as_completed(_futures), **kwargs):
        iftags.append(future.result())
    iftags = [f.result() for f in _futures1]
    texttags = [f.result() for f in _futures2]

    # Export
    iftagsFile = args.iftags.replace(".xlsx", ".translated.xlsx")
    texttagsFile = args.texttags.replace(".xlsx", ".translated.xlsx")

    print("Exporting IF tags to {}".format(iftagsFile))
    print("Exporting text tags to {}".format(texttagsFile))

    to_xlsx(iftags, iftagsFile)
    to_xlsx(texttags, texttagsFile)
